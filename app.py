import streamlit as st
from platform import python_version
import os
import pandas as pd
from datetime import datetime
import openpyxl
from openpyxl.styles import PatternFill, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import Workbook


def tut4(of):
    print('fuck')
    u = of['U']  # assigning a list to columns of input file
    v = of['V']
    w = of['W']
    t = of['T']

    # calculating the avg value sum() returns summation and len() return length
    avgu = [sum(u) / len(u)]
    avgv = [sum(v) / len(v)]
    avgw = [sum(w) / len(w)]

    au = sum(u) / len(u)
    av = sum(v) / len(v)
    aw = sum(w) / len(w)

    u_ = []
    v_ = []
    w_ = []

    for i in u:
        u_.append(i - au)  # pushing the element at back of list
    for i in v:
        v_.append(i - av)
    for i in w:
        w_.append(i - aw)

    # filling the remaining spaces in the column with blank space using extend function
    avgu.extend([''] * (len(u) - 1))
    avgv.extend([''] * (len(u) - 1))
    avgw.extend([''] * (len(u) - 1))

    octs = [1, -1, 2, -2, 3, -3, 4, -4]

    lenc = {}  # dictionary to store longest subsequence length for each octant
    for i in octs:
        lenc[i] = 0

    octants = []  # list to store octants for each case

    col2 = []  # column to store longest subsequence length for each octant
    col4 = []  # column to store octants and time row
    col5 = []  # column to store start time
    col6 = []  # column to store end time

    # loop to determine the octant and to store the octants
    # and update the length of subsequence with maxmimum length each time
    c = 1
    p = 1
    for i in range(len(u_)):
        oc = 1
        if (u_[i] > 0) and (v_[i] > 0) and (w_[i] > 0):
            oc = 1
        elif (u_[i]) < 0 and (v_[i]) > 0 and (w_[i]) > 0:
            oc = 2
        elif (u_[i]) < 0 and (v_[i]) < 0 and (w_[i]) > 0:
            oc = 3
        elif (u_[i]) > 0 and (v_[i]) < 0 and (w_[i]) > 0:
            oc = 4
        elif (u_[i]) > 0 and (v_[i]) > 0 and (w_[i]) < 0:
            oc = -1
        elif (u_[i]) < 0 and (v_[i]) > 0 and (w_[i]) < 0:
            oc = -2
        elif (u_[i]) < 0 and (v_[i]) < 0 and (w_[i]) < 0:
            oc = -3
        elif (u_[i]) > 0 and (v_[i]) < 0 and (w_[i]) < 0:
            oc = -4

        octants.append(oc)
        if i > 0:
            if oc == p:
                c += 1
            else:
                c = 1
        p = oc
        lenc[oc] = max(lenc[oc], c)

    octants.extend([''] * (len(u_) - len(octants)))

    of['         '] = ''

    col1 = []  # column to store the octant values
    for i in octs:
        col1.append(i)

    col1.extend([''] * (len(u_) - len(col1)))
    of['Octant ##'] = col1

    # column to store the longest subsequence length for each octant value
    for i in lenc.values():
        col2.append(i)

    col2.extend([''] * (len(u_) - len(col2)))
    of['Longest Subsequence Length'] = col2

    col3 = []  # column to store count of longest subsequence for each octant value

    maxc = {}  # dictionary to store count of longest subsequence for each octant value
    maxl = {}  # dictionary to store start and end time for each octant's longest subsequence
    for i in octs:
        maxc[i] = 0
        maxl[i] = []

    p = 0
    c = 1
    time = t[0]
    f = 0
    j = 0
    # loop performing required operations to store count of longest subsequence for each octant value
    # and start and end time for each longest subsequence of individual octant
    for i in octants:
        if f == 0:
            time = t[j]
            f = 1
        oc = i
        if oc == p:
            c += 1
        else:
            c = 1
            time = t[j]

        if c == lenc[oc]:
            maxc[oc] += 1
            maxl[oc].append([time, t[j]])
            f = 0

        p = oc
        j = j + 1

    for i in maxc.values():
        col3.append(i)

    col3.extend([''] * (len(u_) - len(col3)))
    of['Count '] = col3

    of['      '] = ''

    # forming the required output columns accordingly from this tut's output
    for i in octs:
        col4.append(i)
        col5.append(lenc[i])
        col6.append(maxc[i])

        col4.append('Time')
        col5.append('From')
        col6.append('To')

        for j in maxl[i]:
            col5.append(j[0])
            col6.append(j[1])

        col4.extend([''] * (len(col6) - len(col4)))

    cnt = len(col6)
    col4.extend([''] * (len(u_) - len(col4)))
    col5.extend([''] * (len(u_) - len(col5)))
    col6.extend([''] * (len(u_) - len(col6)))

    of['Octant ###'] = col4
    of[' Longest Subsequence Length'] = col5
    of['Count  '] = col6
    return cnt


def tut2(of, mod=5000):
    u = of['U']  # assigning list to columns of input file
    v = of['V']
    w = of['W']

    # calculating the avg value sum() returns summation and len() return length
    avgu = [sum(u) / len(u)]
    avgv = [sum(v) / len(v)]
    avgw = [sum(w) / len(w)]

    au = sum(u) / len(u)
    av = sum(v) / len(v)
    aw = sum(w) / len(w)

    u_ = []
    v_ = []
    w_ = []

    for i in u:
        u_.append(i - au)  # pushing the element at back of list
    for i in v:
        v_.append(i - av)
    for i in w:
        w_.append(i - aw)

    # filling the remaining spaces in the column with blank space
    avgu.extend([''] * (len(u) - 1))
    avgv.extend([''] * (len(u) - 1))
    avgw.extend([''] * (len(u) - 1))

    col = ['', 'User Input']
    col.extend([''] * (len(u_) - 2))

    ranges = (len(u_) + mod - 1) // mod

    oc_id = ['Overall Count', 'Mod ' + str(mod)]
    # loop for creating the mod range's column
    for i in range(ranges):
        if i == ranges - 1:
            oc_id.append(str(i * mod) + '-' +
                         str(min((i + 1) * mod - 1, len(u) - 1)))
        else:
            oc_id.append(str(i * mod) + '-' + str((i + 1) * mod - 1))

    octs = [1, -1, 2, -2, 3, -3, 4, -4]
    values = {}
    xx = {}

    ran = 2 + int(len(u_) / mod) + bool(len(u_) // mod) + 14 * \
        (1 + int(len(u_) / mod) + bool(len(u_) // mod))

    num = 2 + int(len(u_) / mod) + bool(len(u_) // mod)

    # initializing the dictionary with value equal to 0 and blank spaces as per requirements
    for i in octs:
        values[i] = [0] * (ran)
        values[i][1] = ''

        for k in range(num - 1):
            for j in range(5):
                values[i][j + 14 * k + num] = ''

        values[i][num + 5] = i
        xx[i] = 0

    octants = []

    # dictionary to store position of octant in columns
    inds = {}
    inds[1] = num + 6
    col[inds[1]] = "From"

    for i in range(len(octs)):
        if i > 0:
            inds[octs[i]] = inds[octs[i - 1]] + 1

    # loop to determine the octant and to store the counts of octants
    # and to store the total transition count
    p = 1
    for i in range(len(u_)):
        oc = 1
        if (u_[i] > 0) and (v_[i] > 0) and (w_[i] > 0):
            oc = 1
        elif (u_[i]) < 0 and (v_[i]) > 0 and (w_[i]) > 0:
            oc = 2
        elif (u_[i]) < 0 and (v_[i]) < 0 and (w_[i]) > 0:
            oc = 3
        elif (u_[i]) > 0 and (v_[i]) < 0 and (w_[i]) > 0:
            oc = 4
        elif (u_[i]) > 0 and (v_[i]) > 0 and (w_[i]) < 0:
            oc = -1
        elif (u_[i]) < 0 and (v_[i]) > 0 and (w_[i]) < 0:
            oc = -2
        elif (u_[i]) < 0 and (v_[i]) < 0 and (w_[i]) < 0:
            oc = -3
        elif (u_[i]) > 0 and (v_[i]) < 0 and (w_[i]) < 0:
            oc = -4

        octants.append(oc)
        values[oc][0] += 1
        values[oc][2 + i // mod] += 1

        if i > 0:
            values[oc][inds[p]] += 1

        p = oc

    # formation and adjustment of octant id column
    for i in range(3):
        oc_id.append('')

    oc_id.append('Overall Transition Count')
    oc_id.append('')
    oc_id.append('Count')
    values[1][num + 4] = "To"

    for i in range(len(octs)):
        oc_id.append(octs[i])

    p = octants[0]
    # loop for individual mod range's transition count and table formation
    for i in range(num - 2):

        for j in range(3):
            oc_id.append('')

        oc_id.append('Mod Transition Count')

        if i == num - 3:
            oc_id.append(str(i * mod) + '-' +
                         str(min((i + 1) * mod - 1, len(u) - 1)))
        else:
            oc_id.append(str(i * mod) + '-' + str((i + 1) * mod - 1))

        values[1][inds[1] + 14 - 2] = "To"
        col[inds[1] + 14] = "From"
        oc_id.append("Octant #")

        for j in octs:
            oc_id.append(j)

        for j in octs:
            values[j][inds[1] + 14 - 1] = j

        for j in range(mod * (i) + 1, min(mod * (i + 1), len(u_) - 1) + 1):
            oc = octants[j]

            if j > 0:
                values[oc][inds[p] + 14] += 1
            p = oc

        for j in octs:
            inds[j] += 14

    oc_id.extend([''] * (len(u_) - len(oc_id)))

    # forming the columns required from this tut's output
    for i in octs:
        values[i].extend([''] * (len(u_) - len(values[i])))

    req1 = []
    req2 = []
    req3 = {}
    for i in octs:
        req3[i] = []

    for i in range(len(oc_id)):
        if i >= num + 4:
            req1.append(col[i])
            req2.append(oc_id[i])
            for j in octs:
                req3[j].append(values[j][i])

    blank = [''] * (len(u_))
    of['ok1'] = blank

    req1.extend([''] * (len(u_) - len(req1)))
    req2.extend([''] * (len(u_) - len(req2)))
    of['ok2' + col[num + 3]] = req1
    of[oc_id[num + 3]] = req2

    i = 3
    for j in octs:
        req3[j].extend([''] * (len(u_) - len(req3[j])))
        of['ok' + str(i) + values[j][num + 3]] = req3[j]
        i = i + 1


cols = {}
cols[1] = 23
cols[-1] = 24
cols[2] = 25
cols[-2] = 26
cols[3] = 27
cols[-3] = 28
cols[4] = 29
cols[-4] = 30


def tut5(of, f, poss, mod=5000):
    # try:
    octant_name_id_mapping = {"1": "Internal outward interaction", "-1": "External outward interaction",
                              "2": "External Ejection", "-2": "Internal Ejection",
                              "3": "External inward interaction", "-3": "Internal inward interaction",
                              "4": "Internal sweep", "-4": "External sweep"}
    
    u = of['U']  # assigning a list to columns of input file
    v = of['V']
    w = of['W']

    # calculating the avg value sum() returns summation and len() returns length
    avgu = [round(sum(u) / len(u), 3)]
    avgv = [round(sum(v) / len(v), 3)]
    avgw = [round(sum(w) / len(w), 3)]

    au = round(sum(u) / len(u), 3)
    av = round(sum(v) / len(v), 3)
    aw = round(sum(w) / len(w), 3)

    u_ = []
    v_ = []
    w_ = []

    for i in u:
        u_.append(round(i - au, 3))  # pushing the element at the end of list using append
    for i in v:
        v_.append(round(i - av, 3))
    for i in w:
        w_.append(round(i - aw, 3))

    # filling the remaining spaces in the column with blank space using extend 
    avgu.extend([''] * (len(u) - 1))
    avgv.extend([''] * (len(u) - 1))
    avgw.extend([''] * (len(u) - 1))

    try:
        of["U Avg"] = avgu  # creating a column in output file
        of["V Avg"] = avgv
        of["W Avg"] = avgw

        of["U'=U-U avg"] = u_
        of["V'=V-V avg"] = v_
        of["W'=W-W avg"] = w_
    except:
        print('Error encountered : Mismatch in length of columns')

    emp = [''] * (len(u_))
    col = ['', 'Mod ' + str(mod)]
    col.extend([''] * (len(u_) - 2))

    ranges = (len(u_) + mod - 1) // mod

    oc_id = ['Overall Count']

    # loop for creating the mod range's column
    for i in range(ranges):
        if i == ranges - 1:
            oc_id.append(str(i * mod) + '-' +
                         str(min((i + 1) * mod - 1, len(u) - 1)))
        else:
            oc_id.append(str(i * mod) + '-' + str((i + 1) * mod - 1))

    octs = [1, -1, 2, -2, 3, -3, 4, -4]
    values = {}  # columns before ranking starts
    ranks = {}  # columns containing ranks
    rank = []
    rank.extend([''] * (len(u_)))
    name = []
    name.extend([''] * (len(u_)))

    ran = 2 + int(len(u_) / mod) + bool(len(u_) // mod) + 14 * \
        (1 + int(len(u_) / mod) + bool(len(u_) // mod))

    num = 1 + int(len(u_) / mod) + bool(len(u_) // mod)

    # initializing the dictionary with value equal to 0 and blank spaces as per requirements
    for i in octs:
        values[i] = [''] * (len(u_))
        values[i] = [0] * (num)
        ranks[i] = [''] * (len(u_))
        ranks[i] = [0] * (num)

    octants = []

    # loop to determine the octant and to store the counts of octants
    # and to store the total transition count

    for i in range(len(u_)):
        oc = 1
        if (u_[i] > 0) and (v_[i] > 0) and (w_[i] > 0):
            oc = 1
        elif (u_[i]) < 0 and (v_[i]) > 0 and (w_[i]) > 0:
            oc = 2
        elif (u_[i]) < 0 and (v_[i]) < 0 and (w_[i]) > 0:
            oc = 3
        elif (u_[i]) > 0 and (v_[i]) < 0 and (w_[i]) > 0:
            oc = 4
        elif (u_[i]) > 0 and (v_[i]) > 0 and (w_[i]) < 0:
            oc = -1
        elif (u_[i]) < 0 and (v_[i]) > 0 and (w_[i]) < 0:
            oc = -2
        elif (u_[i]) < 0 and (v_[i]) < 0 and (w_[i]) < 0:
            oc = -3
        elif (u_[i]) > 0 and (v_[i]) < 0 and (w_[i]) < 0:
            oc = -4

        octants.append(oc)
        values[oc][0] += 1
        values[oc][1 + i // mod] += 1

    # formation and adjustment of octant id column
    for i in range(3):
        oc_id.append('')

    try:
        of['Octant'] = octants
        of[' '] = emp
        of[''] = col  # forming a column in output file
    except:
        print('Error encountered : Mismatch in length of columns')

    oc_id.extend([''] * (len(u_) - len(oc_id)))
    of['Octant ID'] = oc_id

    cnt = {}
    for i in octs:
        cnt[i] = 0
    
    #loop to calculate rank of each octant for each mod range and assignment of the values
    for i in range(num):
        seq = []
        for j in octs:
            seq.append([values[j][i], j])
        seq.sort()

        for j in range(len(seq)):
            ranks[seq[j][1]][i] = 8 - j

        rank[i] = seq[len(seq) - 1][1]
        poss.append([rank[i], i])
        name[i] = octant_name_id_mapping[str(rank[i])]

        if i != 0:
            cnt[rank[i]] = cnt[rank[i]] + 1

    for i in range(1):
        ranks[4].append('')
        ranks[-4].append('')

    ranks[4].append('Octant ID')
    ranks[-4].append('Octant Name')
    rank[num + 1] = 'Count of Rank 1 Mod values'

    # representation of name of each octant and count of rank 1 value
    j = num + 2
    for i in octs:
        ranks[4].append(str(i))
        ranks[-4].append(octant_name_id_mapping[str(i)])
        rank[j] = cnt[i]
        j = j + 1

    #count of each octant in each mod range
    for i in octs:
        values[i].extend([''] * (len(u_) - len(values[i])))
        of[str(i)] = values[i]
    
    #rank columns of each octant
    for i in octs:
        ranks[i].extend([''] * (len(u_) - len(ranks[i])))
        of['Rank Octant ' + str(i)] = ranks[i]

    #forming the output columns
    of['Rank1 Octant ID'] = rank
    of['Rank1 Octant Name'] = name




def tut7(f):
    print('f')
    opdir = 'output'  # forming the output directory if not present
    if not os.path.exists(opdir):
        os.makedirs(opdir)

    # for f in files:
        # if ('input/' + f)[-4:] == 'xlsx':
        # reading the input file
    df = pd.read_excel(f)
    of = df
    print('f')

    u = of['U']
    num = 2 + int(len(u) / mod) + bool(len(u) // mod)

    # calling the required functions
    poss = []
    tut5(of, f, poss, mod)
    tut2(of, mod)
    cnt = int(tut4(of))

    heads = []
    c = 33
    while c <= 43:
        if c != 35:
            heads.append(c)
        c = c+1

    # forming a dataframe in openpyxl from pandas dataframes
    wb = Workbook()
    sheet = wb.active
    for r in dataframe_to_rows(df, index=False, header=True):
        sheet.append(r)

    yellow = "00FFFF00"

    # coloring and bordering the required cells using openpxyl's
    # patternfill and border function
    for i in range(len(poss)):
        c = cols[poss[i][0]]
        r = int(poss[i][1]+2)
        sheet.cell(row=r, column=c).fill = PatternFill(
            patternType="solid", fgColor=yellow)

    for i in range(len(heads)):
        sheet.cell(row=1, column=heads[i]).value = ' '

    c = 14
    borcol = []
    while c <= 32:
        borcol.append(c)
        c = c+1

    black = '000000'
    thin_border = Border(left=Side(style='thin', color=black),
                            right=Side(style='thin', color=black),
                            top=Side(style='thin', color=black),
                            bottom=Side(style='thin', color=black))
    for i in range(len(borcol)):
        for j in range(num):
            sheet.cell(row=j+1, column=borcol[i]).border = thin_border

    for i in range(9):
        sheet.cell(row=num + 2 + i, column=29).border = thin_border
        sheet.cell(row=num + i + 2, column=30).border = thin_border
        sheet.cell(row=num + i + 2, column=31).border = thin_border

    borcol = []
    c = 35
    while c <= 43:
        borcol.append(c)
        c = c+1

    ran = int(len(u)/mod)+bool(len(u) % mod)+1
    row = 0
    for k in range(ran):
        row = row+2
        for i in range(len(borcol)):
            for j in range(9):
                sheet.cell(
                    row=row+1+j, column=borcol[i]).border = thin_border
        row = row+12

    c = 44
    while c <= 46:
        c = c+1
        for i in range(9):
            sheet.cell(row=i + 1, column=c).border = thin_border

    for i in range(cnt+1):
        sheet.cell(row=1 + i, column=49).border = thin_border
        sheet.cell(row=i+1, column=50).border = thin_border
        sheet.cell(row=i+1, column=51).border = thin_border

    # forming the required output file by saving openpyxl dataframe
    wb.save(os.path.join(
            opdir, f.name[0:-4] + '_octant_analysis_mod_' + str(mod) + '.xlsx'))

st.title('Get output file of CS384-2022 tut-7 for free')
f = st.file_uploader('Upload your input file in xlsx format', accept_multiple_files=True)

mod=0
if f is not None:
    mod=int(st.number_input('Please enter mod value'))
    if mod!=0:
        
        if st.button('Compute'):
            for files in f:
                print(files.name)
                tut7(files)        
            
    else:
        st.warning('Mod cannot be zero', icon='🥱')
    